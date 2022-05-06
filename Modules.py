
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Side
from openpyxl.styles.borders import Border, BORDER_THICK
import time


# <---------- Classes ---------------->

class bubble:

    def __init__(self, frame, image, padx=0, pady=0):
        self.my_label = Label(frame, image=image, bg="#F4FCD9")
        self.my_label.pack(padx=padx, pady=pady)


class Door:

    def __init__(self, frame, door_no, close_image, open_car_image, open_goat_image, door_coord=0):
        self.door_no = door_no
        self.close_door_label = Button(frame, image=close_image, bd=0)
        # self.open_door_label = Button(frame, image=open_car_image)
        self.selected = False
        self.door_coord = door_coord

    def __str__(self):
        return f"Door no. {self.door_no}"

    def select_door(self):
        self.selected = True


class Player:
    def __init__(self):
        self.play_time = 0
        self.norm_round = []
        self.switch_round = []

    def save_worksheet(self):

        wb = Workbook()
        ws = wb.active

        # <---- Heading ------>
        # Normal round
        ws["A1"] = "Normal Round"
        ws.merge_cells("A1:B1")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Switching Round
        ws["F1"] = "Switching Round"
        ws.merge_cells("F1:G1")
        ws["F1"].alignment = Alignment(horizontal="center")

        # <-------- Declaring size of cells ------->
        ws.append(["Round", "Result"])
        ws.cell(column=6, row=2, value="Round")
        ws.cell(column=7, row=2, value="Result")

        for x in [["A", "B"], ["F", "G"]]:
            ws.column_dimensions[x[0]].width = 15
            ws.column_dimensions[x[1]].width = 15

        # <-------- Bold Placeholder at top ---------->
        for bold_cell in [1, 2, 6, 7]:
            char = get_column_letter(bold_cell)
            ws[char + "2"].font = Font(bold=True, color="00FFFFFF")
            ws[char + "2"].fill = PatternFill(bgColor="00000000", fill_type="solid")

        # <-------- Appending Data into Cells ----------->
        # For Norm list
        for index, item in enumerate(self.norm_round):
            ws.append([index + 1, item])

        # For Switch list
        for index, item in enumerate(self.switch_round):
            ws.cell(column=6, row=index + 3, value=index + 1)
            ws.cell(column=7, row=index + 3, value=item)

        # <-------- Stylizing --------->
        for col in range(3, len(self.norm_round) + 3):
            # For Norm list
            ws["A" + str(col)].alignment = Alignment(horizontal="center")
            ws["A" + str(col)].border = Border(bottom=Side(border_style=BORDER_THICK, color="000000FF"))

            ws["B" + str(col)].alignment = Alignment(horizontal="center")
            ws["B" + str(col)].font = Font(color="00FF0000")

            # For Switch list
            ws["F" + str(col)].alignment = Alignment(horizontal="center")
            ws["F" + str(col)].border = Border(bottom=Side(border_style=BORDER_THICK, color="000000FF"))

            ws["G" + str(col)].alignment = Alignment(horizontal="center")
            ws["G" + str(col)].font = Font(color="00FF0000")

        # <------- Win Percentage --------->
        # for Normal list
        norm_win_percentage = (self.norm_round.count("Win") * 100) // len(self.norm_round)
        ws.append(["Percentage", norm_win_percentage])

        # For Switch list
        switch_win_percentage = (self.switch_round.count("Win") * 100) // len(self.switch_round)
        ws.cell(column=6, row=len(self.switch_round) + 3, value="Percentage")
        ws.cell(column=7, row=len(self.switch_round) + 3, value=switch_win_percentage)

        # Saving File
        wb.save(f"Test_files/test_result_{str(self.play_time)}.xlsx")

        messagebox.showinfo("Successfully saved",
                            "Your File is saved in File Directory ..\Test_files of this software")


class Animation:
    def __init__(self, root, canvas):
        self.pointer = None
        self.root = root
        self.canvas = canvas

    def door_locater(self, door_no):

        if door_no == 1:
            coord = 250, 1, 250, 21
        elif door_no == 2:
            coord = 650, 1, 650, 21
        else:
            coord = 1050, 1, 1050, 21

        self.pointer = self.canvas.create_line(coord,
                                               width=3, arrow=LAST,
                                               arrowshape=(6, 6, 6),
                                               fill="Green")

    def switch_animation(self):
        circle_coord = [675, 450]

        for direction in range(1, 5):
            for terms in range(25):
                coord = [circle_coord[-2], circle_coord[-1]]
                if direction == 1 or direction == 2:
                    circle_coord.append(coord[0] + 2)
                else:
                    circle_coord.append(coord[0] - 2)

                if direction == 1 or direction == 4:
                    if coord[1] < circle_coord[-3] + 30:
                        circle_coord.append(coord[1] - 2)
                else:
                    if coord[1] < circle_coord[-3] + 30:
                        circle_coord.append(coord[1] + 2)

                self.canvas.create_line(circle_coord, width=5, fill="Green", tags="line_t")
                self.canvas.update()
                time.sleep(0.01)

        time.sleep(1)

        self.canvas.delete("line_t")
